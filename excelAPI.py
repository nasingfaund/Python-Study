import openpyxl
from openpyxl import *

#https://tokmakov.msk.ru/blog/item/71

class ExcelHelper:

	def __init__(self, filePath):
		self.filePath = filePath
		self.workbook = openpyxl.load_workbook(filePath)

	# TODO: проверка на null
	# sheet - индекс листа
	def GetCellValueByAddress(self, sheetIndex, cellAddress):
		sheet = self.GetSheetByIndex(sheetIndex)
		value = sheet[cellAddress].value
		return value


	# Индексы отсчитываются с 1!
	def GetCellValueByIndexes(self, sheetIndex, row, column):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.cell(row, column).value


	# TODO: проверка на null
	def SetCellValueByAddress(self, sheetIndex, cellAddress, value):
		sheet = self.GetSheetByIndex(sheetIndex)
		value = sheet[cellAddress].value = value


	# Индексы отсчитываются с 1!
	def SetCellValueByIndexes(self, sheetIndex, row, column, value):
		sheet = self.GetSheetByIndex(sheetIndex)
		sheet.cell(row, column).value = value


	# TODO: проверка на null
	def GetSheetByIndex(self, sheetIndex):
		sheet = self.workbook.worksheets[sheetIndex]
		return sheet


	# TODO: проверка на null
	def GetSheetByName(self, sheetName):
		sheet = self.workbook[sheetName]
		return sheet


	def RemoveSheetByName(self, sheetName):
		sheet = self.GetSheetByName(sheetName)
		self.workbook.remove(sheet)
		self.SaveFile()


	def RemoveSheetByIndex(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		self.workbook.remove(sheet)
		self.SaveFile()

	
	def GetSheetNameByIndex(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.title


	def GetSheetIndexByName(self, sheetName):
		sheet = self.GetSheetByName(sheetName)
		return self.workbook.index(sheet)


	def CreateNewSheet(self, sheetName):
		sheet = self.workbook.create_sheet(sheetName)
		self.SaveFile()
		return sheet.title


	def SaveFile(self):
		self.workbook.save(self.filePath)


	def GetMaxNotEmptyRowsBySheet(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.max_row


	def GetMaxNotEmptyColumnsBySheet(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.max_column


	def GetFirstNotEmptyRowBySheet(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.min_row


	def GetFirstNotEmptyColumnBySheet(self, sheetIndex):
		sheet = self.GetSheetByIndex(sheetIndex)
		return sheet.min_column

	def IsCellMerged(self, sheetIndex, address):
		sheet = self.GetSheetByIndex(sheetIndex)
		cell = sheet[address]

		for mergedCell in sheet.merged_cells.ranges:
			if (cell.coordinate in mergedCell):
				return True
			return False
  
def GetColumnIndexFromString(columnStr):
	return openpyxl.utils.column_index_from_string(columnStr)

def GetColumnLetterFromIndex(columnIndex):
	return openpyxl.utils.get_column_letter(columnIndex)

def CreateFile(fileName):
	workbook = openpyxl.Workbook()
	workbook.save(fileName)

# using example:
filePath = 'C:/Users/nightblure/Desktop/Производители Минсельхоз.xlsx'

# date, value
#

excelHelper = ExcelHelper(filePath)
text = ''
SHEET_INDEX = 0

for i in range (3, excelHelper.GetMaxNotEmptyRowsBySheet(SHEET_INDEX) + 1):

	value = excelHelper.GetCellValueByIndexes(SHEET_INDEX, i, 12)

	if (not value):
		continue

	date = excelHelper.GetCellValueByIndexes(SHEET_INDEX, i, 1)
	date = date.strftime('%Y-%m-%d')

	query = f"insert into t_goods_raw_data values (3, 7, '{date}', 1, {value});"
	text = text + query + '\n'
	#print(query)

print(text)